import time

import pandas as pd

import openpyxl
from openpyxl import load_workbook

#import mysql.connector

class log_model():
    
    def __init__(self, run_time, datasets, split_train_test, model, params, accuracy, program="program1", split_train_valid=0, metrics=None):
        self.run_time = run_time
        self.program=program
        
        self.datasets=datasets#如果是.__class__.__name__，則輸出奇怪
        self.split_info="train_test_split: {} \n train_valid_split: {}".format(split_train_test, split_train_valid)
        
        self.model=model.__class__.__name__
        self.params=params
        
        parameters=""
        for key in params:
            parameters+="{} : {}\n ".format(key, model.get_params()[key])
        self.params=parameters
        """
        自行設定
        ex. SVC之get_params()有parameter : 
        {'C': 1.0, 'break_ties': False, 'cache_size': 200, 'class_weight': None,
        'coef0': 0.0, 'decision_function_shape': 'ovr', 'degree': 3, 'gamma': 'scale', 
        'kernel': 'rbf', 'max_iter': -1, 'probability': False, 'random_state': 42,
        'shrinking': True, 'tol': 0.001, 'verbose': False}
        """
        self.accuracy=accuracy
        #需要甚麼metrics自行新增
        self.metrics=metrics
        
    def output_params(self, turn_on=False):#turn_on=True, print all info
        if turn_on:
            print("time: ", self.run_time)
            print("program name: ", self.program)
            
            print("datasets: ", self.datasets)
            print("split info: ", self.split_info)
            
            print("model: ", self.model)
            print("params info: ", self.params)
            
            print("accuracy: ", self.accuracy)
            print("metrics: ", self.metrics)
        
    def to_excel(self, record_path, record_sheet):
        
        # 將資料存入字典
        data = {
            #"id by time" : [id_by_time],
            "time": [self.run_time],
            "dataset": [self.datasets],
            "info": [self.split_info],
            "model": [self.model],
            "parameter": [self.params],
            "accuracy": [self.accuracy],
            "metrics": [self.metrics]
        }
        df=pd.DataFrame(data)
        
        try:
            # 載入現有的工作簿
            book = load_workbook(record_path)
            sheet = book[record_sheet]  # 指定目標工作表名稱

            # 找到工作表中第一個空白列
            start_row = sheet.max_row + 1

            # 將 DataFrame 的資料逐行寫入
            for r_idx, row in df.iterrows():
                for c_idx, value in enumerate(row, start=1):
                    sheet.cell(row=start_row + r_idx, column=c_idx, value=value)

            # 儲存檔案
            book.save(record_path)
            print(f"資料已成功輸出至 {record_path}")
            
        except FileNotFoundError:
            # 如果檔案不存在，直接建立新的檔案
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=record_sheet)
            print(f"已成功建立新檔案 {record_path}")
                
        #忘記關閉檔案
        except PermissionError:
            print("請先關閉檔案")
            #exit()
        
    def to_DB(self):
        """
        CREATE DATABASE 
            db_auto_log DEFAULT CHARACTER SET = 'utf8mb4';
        """
        """
        

        # 建立 MySQL 連線
        connection = mysql.connector.connect(
            host="localhost",  # 替換為您的 MySQL 主機
            user="root",  # 替換為您的 MySQL 使用者名稱
            password="0000",  # 替換為您的 MySQL 密碼
            database="your_database"  # 替換為您的資料庫名稱
        )

        cursor = connection.cursor()
        """
        
        # 建立資料表（如果尚未存在）
        #create_table_query = """
        #CREATE TABLE IF NOT EXISTS model_logs (
        #    id INT AUTO_INCREMENT PRIMARY KEY,
        #    time VARCHAR(255),
        #    dataset VARCHAR(255),
        #    info TEXT,
        #    model TEXT,
        #    parameter TEXT,
        #    accuracy FLOAT,
        #    metrics TEXT
        #)
        """
        cursor.execute(create_table_query)

        # 插入資料
        #insert_query = """
        #INSERT INTO model_logs (time, dataset, info, model, parameter, accuracy, metrics)
        #VALUES (%s, %s, %s, %s, %s, %s, %s)
        #"""
        
        """
        data_values = (
            data["run_time"][0],
            data["dataset"][0],
            data["info"][0],
            data["model"][0],
            data["parameter"][0],
            data["accuracy"][0],
            data["metrics"][0]
        )
        cursor.execute(insert_query, data_values)

        # 提交變更並關閉連線
        connection.commit()
        print("資料已成功新增至 MySQL 資料庫。")

        cursor.close()
        connection.close()
        """
        
        pass